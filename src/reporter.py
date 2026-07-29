"""
OBAF — Reporter v2.1
Excel 보고서 (openpyxl) + Telegram 메시지 + Telegram 파일 전송(sendDocument)
"""

import gc
import logging
import os
import requests
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("Reporter")
KST = ZoneInfo("Asia/Seoul")

# 컬러
COLOR = {
    "BUY_HEADER":   "1F4E79", "BUY_ROW":      "DDEEFF",
    "SELL_HEADER":  "7B0000", "SELL_ROW":      "FFE0E0",
    "HOLD_HEADER":  "3A3A3A", "HOLD_ROW":      "F5F5F5",
    "NO_ENTRY_HDR": "4A4A00", "NO_ENTRY_ROW":  "FFFFF0",
    "HEADER_FONT":  "FFFFFF", "TITLE_BG":      "0D1B2A",
    "TITLE_FONT":   "E8F4FD", "MACRO_GOOD":    "C6EFCE",
    "MACRO_WARN":   "FFEB9C", "MACRO_CRIT":    "FFC7CE",
    "BORDER":       "BFBFBF",
}
THIN = Side(border_style="thin", color=COLOR["BORDER"])
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ACTION_META = {
    "BUY_TARGET": {"label": "🚀 BUY",      "row": "BUY_ROW"},
    "SELL_100":   {"label": "🚨 SELL 100%","row": "SELL_ROW"},
    "SELL_50":    {"label": "⚠️ SELL 50%", "row": "SELL_ROW"},
    "NO_ENTRY":   {"label": "⛔ NO ENTRY", "row": "NO_ENTRY_ROW"},
    "HOLD":       {"label": "⏸️ HOLD",      "row": "HOLD_ROW"},
}

def _fill(h): return PatternFill("solid", fgColor=h)
def _bf(c="000000", s=11): return Font(bold=True, color=c, size=s)
def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)


# =============================================================================
# Excel 빌더
# =============================================================================
class ExcelReporter:

    COLS = [
        ("티커",         "ticker",           10),
        ("회사명",       "company",          25),
        ("GICS 섹터",    "gics_sector",      25),
        ("매매 지시",    "action_label",     14),
        ("현재가",       "price",            12),
        ("SMA200",      "sma200",           12),
        ("이격률",       "div_sma200_pct",   12),
        ("RSI",         "rsi",              10),
        ("KDJ-J",       "kdj_j",            10),
        ("DPI 스코어",  "dpi_score",        12),
        ("ICR",         "icr",              10),
        ("D/E",         "de_ratio",         10),
        ("섹터 가중치", "sector_weight",    12),
        ("헤어컷",      "applied_haircut",  10),
        ("유효 가중치", "effective_weight", 12),
        ("판단 근거",   "reasoning",        90),
    ]

    def __init__(self, output_dir: str):
        os.makedirs(output_dir, exist_ok=True)
        self.output_dir = output_dir

    def _hrow(self, ws, row, bg):
        for ci, (h, _, _) in enumerate(self.COLS, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = _fill(bg); c.font = _bf(COLOR["HEADER_FONT"])
            c.alignment = _ctr(); c.border = BOX

    def _write_orders(self, wb, orders, run_time, as_active=False):
        ws = wb.active if as_active else wb.create_sheet("1. 매매 지시서")
        ws.title = "1. 매매 지시서"
        ws.freeze_panes = "A3"
        nc = len(self.COLS)
        ws.merge_cells(f"A1:{get_column_letter(nc)}1")
        t = ws["A1"]
        t.value = f"OBAF 매매 지시서  |  {run_time}  |  총 {len(orders)}개"
        t.fill = _fill(COLOR["TITLE_BG"])
        t.font = Font(bold=True, color=COLOR["TITLE_FONT"], size=14)
        t.alignment = _ctr()
        ws.row_dimensions[1].height = 28
        self._hrow(ws, 2, COLOR["HOLD_HEADER"])
        ws.row_dimensions[2].height = 22

        pri = {"BUY_TARGET":0,"SELL_100":1,"SELL_50":2,"NO_ENTRY":3,"HOLD":4}
        sorted_o = sorted(orders.values(), key=lambda x: pri.get(x["action"], 9))

        for ri, o in enumerate(sorted_o, 3):
            meta = ACTION_META.get(o["action"], ACTION_META["HOLD"])
            bg   = COLOR[meta["row"]]
            data_block   = o.get("_asset_data", {})
            fundamentals = data_block.get("fundamentals", {}) if data_block else {}
            row_data = {
                "ticker":          o.get("ticker",""),
                "company":         o.get("company",""),
                "gics_sector":     o.get("gics_sector",""),
                "action_label":    meta["label"],
                "price":           o.get("price", 0.0),
                "sma200":          o.get("sma200", 0.0),
                "div_sma200_pct":  f"{o.get('div_sma200',0.0):.2%}",
                "rsi":             o.get("rsi", 0.0),
                "kdj_j":           o.get("kdj_j", 0.0),
                "dpi_score":       o.get("dpi_score", 0.0),
                "icr":             fundamentals.get("icr", 0.0),
                "de_ratio":        fundamentals.get("debt_to_equity", 0.0),
                "sector_weight":   o.get("sector_weight", 1.0),
                "applied_haircut": o.get("applied_haircut", 1.0),
                "effective_weight":o.get("effective_weight", 1.0),
                "reasoning":       o.get("reasoning",""),
            }
            for ci, (_, key, _) in enumerate(self.COLS, 1):
                cell = ws.cell(row=ri, column=ci, value=row_data[key])
                cell.fill = _fill(bg); cell.border = BOX
                cell.alignment = _lft() if ci == nc else _ctr()
            ws.row_dimensions[ri].height = 18

        for ci, (_, _, w) in enumerate(self.COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ------------------------------------------------------------------
    # Phase -1 월봉 RSI GATE 통과 종목 — 메인 리포트 시트
    # ------------------------------------------------------------------
    def _write_monthly_rsi_survivors(
        self, wb, gate_details: dict, threshold: float,
        core_entry_min_rsi: float = 40, core_entry_max_rsi: float = 50
    ):
        """
        Phase -1(월봉 RSI 사전 필터) 통과 종목을 메인 리포트의 첫 시트에 표시한다.
        (메인 리포트는 통과 종목이 1개 이상 있을 때만 생성되므로 이 시트는
        항상 PASS 종목만 포함한다. 탈락 종목 전체 검증은 별도의
        OBAF_MonthlyRSIGate_*.xlsx 파일에서 확인 가능하다.)
        """
        ws = wb.active
        ws.title = "0. 월봉RSI 통과종목"

        passed = sorted(
            [d for d in gate_details.values() if d.get("pass")],
            key=lambda d: d.get("monthly_rsi") if d.get("monthly_rsi") is not None else 999.0
        )
        core_entry_hits = [d for d in passed if d.get("in_core_entry_band")]
        band_label = f"CORE ENTRY 구간({core_entry_min_rsi:g}~{core_entry_max_rsi:g})"

        headers = ["티커", "회사명", "GICS 섹터", "월봉 RSI", "일봉 RSI", band_label]
        widths  = [10, 25, 25, 12, 12, 40]
        nc = len(headers)

        ws.merge_cells(f"A1:{get_column_letter(nc)}1")
        t = ws["A1"]
        t.value = (
            f"OBAF Phase -1 월봉 RSI GATE 통과 종목  |  기준: 월봉 RSI < {threshold}  |  "
            f"총 {len(passed)}개 통과  |  이 중 {band_label} 해당: {len(core_entry_hits)}개"
        )
        t.fill = _fill(COLOR["TITLE_BG"])
        t.font = Font(bold=True, color=COLOR["TITLE_FONT"], size=13)
        t.alignment = _ctr()
        ws.row_dimensions[1].height = 30

        ws.freeze_panes = "A3"
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.fill = _fill(COLOR["HOLD_HEADER"]); c.font = _bf(COLOR["HEADER_FONT"])
            c.alignment = _ctr(); c.border = BOX
        ws.row_dimensions[2].height = 20

        for ri, d in enumerate(passed, 3):
            in_band = d.get("in_core_entry_band")
            bg = COLOR["MACRO_GOOD"] if in_band else COLOR["BUY_ROW"]
            monthly_rsi_val = d.get("monthly_rsi")
            daily_rsi_val   = d.get("daily_rsi")
            band_cell_val   = "해당" if in_band is True else ("미해당" if in_band is False else "N/A")
            row_vals = [
                d.get("ticker", ""),
                d.get("company", ""),
                d.get("gics_sector", ""),
                monthly_rsi_val if monthly_rsi_val is not None else "N/A",
                daily_rsi_val if daily_rsi_val is not None else "N/A",
                band_cell_val,
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = _fill(bg); c.border = BOX
                c.alignment = _ctr()
            ws.row_dimensions[ri].height = 18

        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _write_macro(self, wb, macro_data, dash_score):
        ws = wb.create_sheet("2. 매크로")
        headers = ["지표","현재값","임계값","상태"]
        widths  = [25, 15, 15, 20]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = _fill(COLOR["TITLE_BG"]); c.font = _bf(COLOR["TITLE_FONT"])
            c.alignment = _ctr(); c.border = BOX
            ws.column_dimensions[get_column_letter(ci)].width = w

        rows = [
            ("US10Y (%)",         macro_data.get("US10Y",    0.0), 4.5,  True),
            ("WTI ($/bbl)",        macro_data.get("CrudeOil", 0.0), 95.0, True),
            ("지정학 리스크",      macro_data.get("GeoIndex", 0.0), 75.0, True),
            ("Risk-On DashScore",  dash_score,                      40.0, False),
        ]
        for ri, (label, val, thr, lower_good) in enumerate(rows, 2):
            is_ok = (val < thr) if lower_good else (val >= thr)
            bg    = COLOR["MACRO_GOOD"] if is_ok else COLOR["MACRO_WARN"]
            status = "✅ 정상" if is_ok else "⚠️ 경고"
            for ci, v in enumerate([label, val, thr, status], 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = _fill(bg); c.alignment = _ctr(); c.border = BOX
            ws.row_dimensions[ri].height = 20

    def _write_pipeline(self, wb, stats):
        ws = wb.create_sheet("3. 파이프라인")
        headers = ["단계","진입","통과","탈락","생존율"]
        phases  = [
            ("Phase -1 — 월봉 RSI GATE (1차 관문)", "phase_m1_in","phase_m1_out"),
            ("Phase 0 — VETO GATE",           "phase0_in","phase0_out"),
            ("Phase 1 — B_NECK_STRAT",         "phase1_in","phase1_out"),
            ("Phase 2 — VALUATION & SURVIVAL", "phase2_in","phase2_out"),
            ("Phase 3 — TECH EXECUTION",       "phase3_in","phase3_out"),
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = _fill(COLOR["TITLE_BG"]); c.font = _bf(COLOR["TITLE_FONT"])
            c.alignment = _ctr(); c.border = BOX

        for ri, (name, ki, ko) in enumerate(phases, 2):
            p_in  = stats.get(ki, 0)
            p_out = stats.get(ko, 0)
            elim  = p_in - p_out
            sur   = f"{p_out/p_in:.1%}" if p_in > 0 else "N/A"
            for ci, v in enumerate([name, p_in, p_out, elim, sur], 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.alignment = _ctr(); c.border = BOX
            ws.row_dimensions[ri].height = 20
        for ci, w in enumerate([38,12,12,12,12], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ------------------------------------------------------------------
    # Phase -1 월봉 RSI 게이트 — 검증용 Excel
    # ------------------------------------------------------------------
    def generate_monthly_rsi_gate_report(
        self, gate_details: dict, threshold: float,
        core_entry_min_rsi: float = 40, core_entry_max_rsi: float = 50
    ) -> str:
        """
        Phase -1(월봉 RSI 사전 필터) 결과를 검증할 수 있는 Excel 생성.
        전체 종목 수가 많으므로:
          - 통과(PASS) 종목이 1개 이상 있으면 → 통과 종목만 표시
          - 통과 종목이 0개이면 → 탈락(전체) 종목을 표시하여 원인을 검증 가능하게 함
        필터링 결과가 0건이어도 항상 생성된다.

        통과 종목에 한해 일봉 RSI-14(daily_rsi)와 Phase 3 CORE ENTRY RSI 구간
        (기본 40~50) 해당 여부(in_core_entry_band)를 함께 표기해, 월봉 게이트
        통과 종목이 실제로 Phase 3 CORE ENTRY 조건까지 도달할 가능성이 있는지
        바로 확인할 수 있게 한다. (월봉 RSI와 일봉 RSI는 서로 다른 시계열
        지표이므로 상충되지 않으며, 이 컬럼은 그 사실을 데이터로 보여주기 위한
        교차 검증용이다.)
        """
        now      = datetime.now(tz=KST)
        run_time = now.strftime("%Y-%m-%d %H:%M KST")
        filename = f"OBAF_MonthlyRSIGate_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        all_rows = list(gate_details.values())
        passed   = [d for d in all_rows if d.get("pass")]
        rejected = [d for d in all_rows if not d.get("pass")]
        core_entry_hits = [d for d in passed if d.get("in_core_entry_band")]

        band_label = f"CORE ENTRY 구간({core_entry_min_rsi:g}~{core_entry_max_rsi:g})"
        headers = ["티커", "회사명", "GICS 섹터", "월봉 RSI", "일봉 RSI", band_label, "판정", "사유"]
        widths  = [10, 25, 25, 12, 12, 20, 10, 70]

        wb = Workbook()
        ws = wb.active

        def _rsi_sort_key(d, default):
            v = d.get("monthly_rsi")
            return v if v is not None else default

        if passed:
            ws.title = "월봉RSI 통과종목"
            mode_label = (
                f"통과 종목만 표시 ({len(all_rows)}개 중 {len(passed)}개 통과 → 통과분만 노출)  |  "
                f"통과 종목 중 일봉 RSI가 {band_label}에 해당: {len(core_entry_hits)}개"
            )
            rows = sorted(passed, key=lambda d: _rsi_sort_key(d, 999.0))
        else:
            ws.title = "월봉RSI 탈락종목(전체)"
            mode_label = "통과 종목 0개 → 탈락 전체 종목 표시 (검증용)"
            rows = sorted(rejected, key=lambda d: -_rsi_sort_key(d, -1.0))

        nc = len(headers)
        ws.merge_cells(f"A1:{get_column_letter(nc)}1")
        t = ws["A1"]
        t.value = (
            f"OBAF Phase -1 월봉 RSI 게이트  |  {run_time}  |  기준: 월봉 RSI < {threshold} 만 통과  |  "
            f"입력 {len(all_rows)}개 → 통과 {len(passed)}개 / 탈락 {len(rejected)}개  |  {mode_label}"
        )
        t.fill = _fill(COLOR["TITLE_BG"])
        t.font = Font(bold=True, color=COLOR["TITLE_FONT"], size=12)
        t.alignment = _ctr()
        ws.row_dimensions[1].height = 32

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.fill = _fill(COLOR["HOLD_HEADER"]); c.font = _bf(COLOR["HEADER_FONT"])
            c.alignment = _ctr(); c.border = BOX
        ws.row_dimensions[2].height = 20

        for ri, d in enumerate(rows, 3):
            is_pass    = bool(d.get("pass"))
            in_band    = d.get("in_core_entry_band")
            if is_pass and in_band:
                bg = COLOR["MACRO_GOOD"]           # 월봉+일봉 조건 동시 충족 후보 → 강조
            elif is_pass:
                bg = COLOR["BUY_ROW"]
            else:
                bg = COLOR["SELL_ROW"]

            monthly_rsi_val = d.get("monthly_rsi")
            daily_rsi_val   = d.get("daily_rsi")
            band_cell_val   = "해당" if in_band is True else ("미해당" if in_band is False else "N/A")

            row_vals = [
                d.get("ticker", ""),
                d.get("company", ""),
                d.get("gics_sector", ""),
                monthly_rsi_val if monthly_rsi_val is not None else "N/A",
                daily_rsi_val if daily_rsi_val is not None else "N/A",
                band_cell_val,
                "PASS" if is_pass else "REJECT",
                d.get("reason", ""),
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = _fill(bg); c.border = BOX
                c.alignment = _lft() if ci == nc else _ctr()
            ws.row_dimensions[ri].height = 18

        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        wb.save(filepath)
        logger.info(f"월봉 RSI 게이트 검증 Excel 저장: {filepath}")
        gc.collect()
        return filepath

    def generate(
        self, execution_orders, macro_data, dash_score, pipeline_stats,
        gate_details: dict | None = None, monthly_rsi_threshold: float = 30,
        core_entry_min_rsi: float = 40, core_entry_max_rsi: float = 50
    ) -> str:
        now      = datetime.now(tz=KST)
        run_time = now.strftime("%Y-%m-%d %H:%M KST")
        filename = f"OBAF_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        if gate_details:
            # 0번 시트: Phase -1 월봉 RSI GATE 통과 종목
            self._write_monthly_rsi_survivors(
                wb, gate_details, monthly_rsi_threshold,
                core_entry_min_rsi, core_entry_max_rsi
            )
            self._write_orders(wb, execution_orders, run_time, as_active=False)
        else:
            # gate_details 미전달 시 기존과 동일하게 매매 지시서를 첫 시트로 사용
            self._write_orders(wb, execution_orders, run_time, as_active=True)
        self._write_macro(wb, macro_data, dash_score)
        self._write_pipeline(wb, pipeline_stats)
        wb.save(filepath)
        logger.info(f"Excel 저장: {filepath}")
        gc.collect()
        return filepath


# =============================================================================
# Telegram 디스패처 v2.1 — sendMessage + sendDocument
# =============================================================================
class TelegramDispatcher:

    MSG_URL = "https://api.telegram.org/bot{token}/sendMessage"
    DOC_URL = "https://api.telegram.org/bot{token}/sendDocument"

    def __init__(self, bot_token: str, chat_id: str):
        self.bot_token = bot_token
        self.chat_id   = chat_id

    def _ok(self) -> bool:
        return bool(self.bot_token and self.chat_id)

    def _send_text(self, text: str) -> bool:
        if not self._ok():
            logger.warning("Telegram 미설정 → 스킵")
            return False
        try:
            url  = self.MSG_URL.format(token=self.bot_token)
            resp = requests.post(
                url,
                json={"chat_id": self.chat_id, "text": text, "parse_mode": "HTML"},
                timeout=15
            )
            resp.raise_for_status()
            return True
        except Exception as e:
            logger.error(f"Telegram 텍스트 발송 실패: {e}")
            return False

    def _send_document(self, filepath: str, caption: str = "") -> bool:
        """Excel 파일을 Telegram으로 직접 전송 (sendDocument API)"""
        if not self._ok():
            return False
        if not os.path.exists(filepath):
            logger.error(f"파일 없음: {filepath}")
            return False
        try:
            url = self.DOC_URL.format(token=self.bot_token)
            with open(filepath, "rb") as f:
                resp = requests.post(
                    url,
                    data={"chat_id": self.chat_id, "caption": caption},
                    files={"document": (os.path.basename(filepath), f,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    timeout=60
                )
            resp.raise_for_status()
            logger.info(f"Telegram Excel 전송 완료: {os.path.basename(filepath)}")
            return True
        except Exception as e:
            logger.error(f"Telegram 파일 전송 실패: {e}")
            return False

    def dispatch_summary(
        self,
        execution_orders: dict,
        macro_data: dict,
        dash_score: float,
        pipeline_stats: dict,
        excel_path: str
    ):
        now_str = datetime.now(tz=KST).strftime("%Y-%m-%d %H:%M KST")

        buy_orders  = [o for o in execution_orders.values() if o["action"] == "BUY_TARGET"]
        sell_100_orders = [o for o in execution_orders.values() if o["action"] == "SELL_100"]
        sell_50_orders = [o for o in execution_orders.values() if o["action"] == "SELL_50"]
        sell_orders = sell_100_orders + sell_50_orders
        hold_count  = sum(1 for o in execution_orders.values() if o["action"] == "HOLD")
        fomo_count  = sum(1 for o in execution_orders.values() if o["action"] == "NO_ENTRY")

        us10y = macro_data.get("US10Y",    0.0)
        oil   = macro_data.get("CrudeOil", 0.0)
        geo   = macro_data.get("GeoIndex", 0.0)

        macro_icon = "🟢" if dash_score >= 60 else ("🟡" if dash_score >= 40 else "🔴")

        lines = [
            f"<b>� OBAF 시스템 리포트</b>  <i>{now_str}</i>",
            "",
            "<b>【매크로 대시보드】</b>",
            f"  {macro_icon} Risk-On DashScore: <b>{dash_score:.1f}</b>",
            f"  📈 US10Y: {us10y:.2f}%",
            f"  🛢 WTI: ${oil:.2f}",
            f"  🌐 지정학 리스크: {geo:.1f}",
            "",
            "<b>【파이프라인 결과】</b>",
            f"  유니버스: {pipeline_stats.get('phase0_in', 0)}개",
            f"  Phase0→1: {pipeline_stats.get('phase0_out', 0)}개",
            f"  Phase1→2: {pipeline_stats.get('phase1_out', 0)}개",
            f"  Phase2→3: {pipeline_stats.get('phase2_out', 0)}개",
            "",
            "<b>【매매 지시 요약】</b>",
            f"  🚀 BUY     : {len(buy_orders)}개",
            f"  🔴 SELL100 : {len(sell_100_orders)}개",
            f"  🟠 SELL50  : {len(sell_50_orders)}개",
            f"  ⏸ HOLD    : {hold_count}개",
            f"  ⛔ NO ENTRY: {fomo_count}개",
        ]

        if buy_orders:
            lines += ["", "<b>🚀 BUY 종목</b>"]
            for o in buy_orders:
                lines.append(
                    f"  • <b>{o['ticker']}</b> ({o['gics_sector']}) "
                    f"${o['price']:.2f} | RSI:{o['rsi']:.1f} | DPI:{o['dpi_score']:.3f}"
                )

        #if sell_orders:
        #    lines += ["", "<b>🔴 SELL 종목</b>"]
        #    for o in sell_100_orders:
        #        lbl = "100%" if o["action"] == "SELL_100" else "50%"
        #        lines.append(f"  • <b>{o['ticker']}</b> SELL {lbl}")

        # sell 100% 종목만 출력
        if sell_100_orders:
            lines += ["", "<b>🔴 SELL 100% 종목</b>"]
            for o in sell_100_orders:
                lines.append(f"  • <b>{o['ticker']}</b> SELL 100%")
        
        msg = "\n".join(lines)

        # 텍스트 발송 (4000자 분할)
        MAX = 4000
        while msg:
            self._send_text(msg[:MAX])
            msg = msg[MAX:]

        # ── Excel 파일 직접 전송 (핵심 수정) ─────────────────────
        caption = f"OBAF Excel 리포트 {now_str}"
        self._send_document(excel_path, caption)
